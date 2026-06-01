#!/usr/bin/env python3
"""
WiFi 异常自愈能力测试脚本
===========================
每轮重启设备，检测开机后 WiFi 是否正常。
若异常则依次尝试"开关 WiFi"和"飞行模式切换"两种恢复方案。
Android 版本通用，命令多级降级，uiautomator2 可选。

用法:
  一键启动: 双击 run.bat
  python wifi_self_healing_test.py --cycles 50
  python wifi_self_healing_test.py --cycles 100 --ssid "MyWiFi" --password "12345678"
  python wifi_self_healing_test.py --cycles 20 --boot-wait 45
"""

import argparse
import codecs
import json
import os
import random
import re
import subprocess
import sys
import time
from dataclasses import dataclass, field
from datetime import datetime
from typing import Dict, List, Optional, Tuple

# ============================================================
# 常量
# ============================================================

WIFI_TOGGLE_OFF_WAIT = 3
WIFI_TOGGLE_ON_WAIT = 5
AIRPLANE_ON_WAIT = 3
AIRPLANE_OFF_WAIT = 8
SCAN_WAIT = 5
CONNECT_WAIT = 5
BOOT_POLL_INTERVAL = 2          # 每 2 秒检测一次是否完成启动
BOOT_TIMEOUT = 300              # 启动超时 5 分钟

DEFAULT_CYCLES = 50
DEFAULT_BOOT_WAIT = 15          # 系统初始化额外等待秒数
DEFAULT_OUTPUT = "./reports"
DEFAULT_SETTINGS_INTENT = "android.settings.WIFI_SETTINGS"
DEFAULT_RECOVERY_MODE = "degrade"

RECOVERY_MODES = {
    "degrade":  "降级链（方案一失败→方案二）",
    "both":     "方案一+方案二都测（同一故障类型）",
}

# 命令探测结果: "svc" / "cmd_wifi" / "none"
_wifi_cmd_mode: Optional[str] = None


# ============================================================
# 类 1: AdbHelper — ADB 命令封装
# ============================================================

class AdbHelper:
    """ADB shell 命令封装，含设备重启、启动等待、命令降级。"""

    @staticmethod
    def _adb_prefix() -> List[str]:
        return ["adb"]

    @classmethod
    def run(cls, shell_cmd: str, timeout: int = 15, root: bool = False) -> Tuple[int, str, str]:
        """执行 adb shell 命令，返回 (exit_code, stdout, stderr)。"""
        if root:
            cls._ensure_root()
        full_cmd = cls._adb_prefix() + ["shell", shell_cmd]
        try:
            proc = subprocess.run(
                full_cmd, capture_output=True, text=True,
                encoding="utf-8", errors="replace", timeout=timeout,
            )
            return proc.returncode, (proc.stdout or "").strip(), (proc.stderr or "").strip()
        except subprocess.TimeoutExpired:
            return -1, "", f"命令超时 ({timeout}s): {shell_cmd}"
        except FileNotFoundError:
            return -1, "", "adb 未找到，请确认 ADB 已安装并在 PATH 中"

    @classmethod
    def run_raw(cls, args: List[str], timeout: int = 15) -> Tuple[int, str, str]:
        """执行原始 adb 命令（非 shell 模式）。"""
        full_cmd = cls._adb_prefix() + args
        try:
            proc = subprocess.run(
                full_cmd, capture_output=True, text=True,
                encoding="utf-8", errors="replace", timeout=timeout,
            )
            return proc.returncode, (proc.stdout or "").strip(), (proc.stderr or "").strip()
        except subprocess.TimeoutExpired:
            return -1, "", f"命令超时 ({timeout}s): {' '.join(args)}"
        except FileNotFoundError:
            return -1, "", "adb 未找到，请确认 ADB 已安装并在 PATH 中"

    @classmethod
    def get_device_info(cls) -> str:
        brand = cls.run("getprop ro.product.brand")[1] or "unknown"
        model = cls.run("getprop ro.product.model")[1] or "unknown"
        sdk = cls.run("getprop ro.build.version.sdk")[1] or "unknown"
        return f"{brand}/{model}/API {sdk}"

    @classmethod
    def check_device(cls) -> bool:
        ret, out, _ = cls.run_raw(["devices"])
        if ret != 0:
            return False
        for line in out.split("\n"):
            if "device" in line and "offline" not in line and "unauthorized" not in line and "List" not in line:
                return True
        return False

    @classmethod
    def _ensure_root(cls) -> bool:
        ret, out, _ = cls.run("id -u")
        if ret == 0 and out.strip() == "0":
            return True
        cls.run_raw(["root"], timeout=5)
        return True

    # ---- 设备重启与启动等待 ----

    @classmethod
    def reboot(cls) -> bool:
        """重启设备。"""
        ret, _, _ = cls.run_raw(["reboot"], timeout=10)
        return ret == 0

    @classmethod
    def wait_for_device(cls, timeout: int = BOOT_TIMEOUT) -> bool:
        """等待设备从离线状态恢复为 online。若 unauthorized 则提示用户授权。"""
        elapsed = 0
        warned_unauth = False
        while elapsed < timeout:
            time.sleep(BOOT_POLL_INTERVAL)
            elapsed += BOOT_POLL_INTERVAL
            ret, out, _ = cls.run_raw(["devices"])
            if ret != 0:
                continue

            # 检查 unauthorized 状态
            if "unauthorized" in out:
                if not warned_unauth:
                    print("\n  ╔══════════════════════════════════════════╗")
                    print("  ║  设备重启后 USB 调试授权丢失            ║")
                    print("  ║  请在设备屏幕上点击「允许」USB 调试     ║")
                    print("  ║  并勾选「一律允许」以持久化授权          ║")
                    print("  ╚══════════════════════════════════════════╝")
                    warned_unauth = True
                continue

            for line in out.split("\n"):
                if "device" in line and "offline" not in line and "List" not in line:
                    return True
        return False

    @classmethod
    def wait_for_boot_completed(cls, timeout: int = BOOT_TIMEOUT) -> bool:
        """等待 sys.boot_completed=1。"""
        elapsed = 0
        while elapsed < timeout:
            time.sleep(BOOT_POLL_INTERVAL)
            elapsed += BOOT_POLL_INTERVAL
            ret, out, _ = cls.run("getprop sys.boot_completed")
            if ret == 0 and out.strip() == "1":
                return True
        return False


# ============================================================
# 数据类: 测试记录
# ============================================================

@dataclass
class CycleRecord:
    """单轮测试记录，降级链：同一故障下方案一→方案二。"""
    cycle: int
    timestamp: str

    # 时间戳（精确到秒）
    boot_completed_time: str = ""    # Android 开机完成时间
    wifi_ready_time: str = ""        # WiFi 列表加载完成时间
    wifi_init_duration_s: float = 0.0  # WiFi 初始化耗时(秒)

    # 基线（注入前采集）
    baseline_ap: int = 0             # 注入前 AP 数量
    baseline_connected: bool = False # 注入前是否已连接热点

    # 故障注入
    fault_ok: bool = False           # 故障注入是否成功
    fault_type: str = ""             # 注入的故障类型标识

    # 方案一：开关 WiFi
    m1_recovered: bool = False       # 恢复是否成功（AP >= 基线80%）
    m1_time_ms: int = 0              # 恢复耗时
    m1_scan_after: int = 0           # 恢复后 AP 数量
    m1_toggle_ok: bool = False       # 恢复后开关是否正常
    m1_connect_ok: bool = False      # 恢复后连接是否成功
    m1_detail: str = ""

    # 方案二：飞行模式（仅方案一失败时执行）
    m2_skipped: bool = True          # 方案一成功则跳过方案二
    m2_recovered: bool = False
    m2_time_ms: int = 0
    m2_scan_after: int = 0
    m2_toggle_ok: bool = False
    m2_connect_ok: bool = False
    m2_detail: str = ""

    # 最终状态
    reboot_ok: bool = False          # 关机/重启命令是否成功
    scan_empty: bool = True          # WiFi列表是否为空（最终）
    log_path: str = ""               # 本轮日志路径

    @property
    def success(self) -> bool:
        """方案一成功 或 (方案一失败但方案二成功)"""
        return self.m1_recovered or (not self.m1_recovered and self.m2_recovered)

    @property
    def final_scan_count(self) -> int:
        """最终恢复后 AP 数量。"""
        if self.m2_skipped:
            return self.m1_scan_after
        return self.m2_scan_after if self.m2_recovered else self.m1_scan_after

    @property
    def final_connect_ok(self) -> bool:
        """最终连接是否成功。"""
        if self.m2_skipped:
            return self.m1_connect_ok
        return self.m2_connect_ok if self.m2_recovered else self.m1_connect_ok

    @property
    def failure_type(self) -> str:
        """失败类型（用于报告）。"""
        if self.success:
            return ""
        if not self.reboot_ok:
            return "重启失败"
        if not self.fault_ok:
            return "故障注入失败"
        if not self.m1_recovered and not self.m2_recovered:
            return "双双失败"
        return "其他"

    @property
    def failure_detail(self) -> str:
        """失败详情。"""
        if self.success:
            return ""
        parts = []
        if not self.reboot_ok:
            parts.append("重启命令失败")
        if not self.fault_ok:
            parts.append("故障注入未生效")
        if not self.m1_recovered and self.m1_detail:
            parts.append(f"方案一: {self.m1_detail}")
        if not self.m2_skipped and not self.m2_recovered and self.m2_detail:
            parts.append(f"方案二: {self.m2_detail}")
        return "; ".join(parts)

    @property
    def m1_recovery_pct(self) -> float:
        """方案一恢复率 (vs 基线)。"""
        if self.baseline_ap <= 0:
            return 100.0 if self.m1_scan_after > 0 else 0.0
        return min(100.0, self.m1_scan_after / self.baseline_ap * 100)

    @property
    def m2_recovery_pct(self) -> float:
        """方案二恢复率 (vs 基线)。"""
        if self.baseline_ap <= 0:
            return 100.0 if self.m2_scan_after > 0 else 0.0
        return min(100.0, self.m2_scan_after / self.baseline_ap * 100)

    def to_dict(self) -> dict:
        return {
            "cycle": self.cycle,
            "timestamp": self.timestamp,
            "boot_completed_time": self.boot_completed_time,
            "wifi_ready_time": self.wifi_ready_time,
            "wifi_init_duration_s": round(self.wifi_init_duration_s, 3),
            "reboot_ok": self.reboot_ok,
            "scan_empty": self.scan_empty,
            "success": self.success,
            "failure_type": self.failure_type,
            "failure_detail": self.failure_detail,
            "log_path": self.log_path,
            "baseline": {"ap_count": self.baseline_ap, "connected": self.baseline_connected},
            "fault": {"ok": self.fault_ok, "type": self.fault_type},
            "method1_wifi_toggle": {
                "recovered": self.m1_recovered, "time_ms": self.m1_time_ms,
                "scan_after": self.m1_scan_after, "recovery_pct": round(self.m1_recovery_pct, 1),
                "toggle_ok": self.m1_toggle_ok, "connect_ok": self.m1_connect_ok,
                "detail": self.m1_detail,
            },
            "method2_airplane_mode": {
                "skipped": self.m2_skipped,
                "recovered": self.m2_recovered, "time_ms": self.m2_time_ms,
                "scan_after": self.m2_scan_after, "recovery_pct": round(self.m2_recovery_pct, 1),
                "toggle_ok": self.m2_toggle_ok, "connect_ok": self.m2_connect_ok,
                "detail": self.m2_detail,
            },
        }


# ============================================================
# 类 2: WifiDetector — WiFi 状态检测
# ============================================================

class WifiDetector:
    """检测 WiFi 健康状态：开关状态、扫描结果、连接能力。"""

    # ---- WiFi 服务就绪等待 ----

    @classmethod
    def wait_for_wifi_ready(cls, max_wait: int = 120, poll_interval: int = 3) -> bool:
        """
        等待 WiFi 服务就绪。
        在 boot_completed 之后 WiFi HAL/wpa_supplicant 可能还没初始化完毕，
        直接调用 dumpsys wifi 会阻塞。这里用快速命令轮询。
        返回 True 表示就绪，False 表示超时。
        """
        print(f"  等待 WiFi 服务就绪...")
        elapsed = 0
        while elapsed < max_wait:
            # 用轻量命令探测 WiFi 服务是否可响应
            rc, out, err = AdbHelper.run("cmd wifi status", timeout=8)
            if rc == 0 and out:
                # 不管内容是什么，有响应说明服务可用
                if "disabled" not in out.lower() or "enabled" in out.lower():
                    print(f"  WiFi 服务就绪 ({elapsed}s)")
                    return True
                # 即使显示 disabled 也是有效响应
                print(f"  WiFi 服务就绪 ({elapsed}s)")
                return True
            if "not found" in (out + err).lower() or "No such command" in (out + err):
                # cmd wifi 不可用，尝试 svc
                rc2, out2, _ = AdbHelper.run("svc wifi", timeout=5)
                if rc2 >= 0 and "not found" not in out2.lower():
                    print(f"  WiFi 服务就绪 (svc, {elapsed}s)")
                    return True

            sys.stdout.write(f"\r  等待 WiFi 服务... {elapsed}s / {max_wait}s")
            sys.stdout.flush()
            time.sleep(poll_interval)
            elapsed += poll_interval

        print(f"\n  [WARN] WiFi 服务等待超时 ({max_wait}s)，尝试继续...")
        return False

    # ---- 基础检测 ----

    @staticmethod
    def is_wifi_enabled() -> bool:
        """检查 WiFi 是否已开启。"""
        _, out, _ = AdbHelper.run("dumpsys wifi | grep -i 'Wi-Fi is'", timeout=10)
        return "enabled" in out.lower()

    @staticmethod
    def enable_wifi() -> bool:
        """尝试开启 WiFi，svc 优先。"""
        rc, _, _ = AdbHelper.run("svc wifi enable", timeout=10)
        if rc != 0:
            AdbHelper.run("cmd wifi set-wifi-enabled enabled", timeout=10)
        time.sleep(SCAN_WAIT)
        return WifiDetector.is_wifi_enabled()

    @staticmethod
    def get_scan_results() -> List[str]:
        """
        获取 WiFi 扫描结果中的 SSID 列表。
        多重路径 + 多格式匹配，适配不同 Android 版本和 ROM。
        """
        out = ""

        # 路径 1: cmd wifi list-scan-results (Android 10+)
        _, raw, _ = AdbHelper.run("cmd wifi list-scan-results", timeout=15)
        if raw and "No such command" not in raw and "not found" not in raw.lower():
            out = raw
        else:
            # 路径 2: dumpsys wifi 全文（部分设备 grep 不可用）
            _, out, _ = AdbHelper.run("dumpsys wifi", timeout=15)

        ssids = WifiDetector._parse_ssids(out)

        # 如果仍然为空，尝试 wpa_cli
        if not ssids:
            _, out2, _ = AdbHelper.run("wpa_cli -i wlan0 scan_results 2>/dev/null", timeout=10)
            if out2:
                ssids = WifiDetector._parse_ssids(out2)
                ssids = [s for s in ssids if s]  # 过滤空行

        # 诊断：打印部分原始输出
        if not ssids and out:
            preview = out[:800].replace("\n", "\\n")
            print(f"    [诊断] 扫描命令输出 (前800字符): {preview}")

        return ssids

    @staticmethod
    def _parse_ssids(text: str) -> List[str]:
        """从文本中提取 SSID，支持多种格式。"""
        ssids = []
        for line in text.split("\n"):
            line = line.strip()
            if not line or len(line) < 2:
                continue

            # 格式 1: SSID: "xxx" 或 SSID: xxx
            m = re.search(r'SSID[:\s=]+"?([^",\t]+)"?', line, re.IGNORECASE)
            if m:
                ssid = m.group(1).strip()
                # 过滤明显不是 SSID 的内容
                if ssid and len(ssid) >= 1 and ssid not in (
                    "null", "<unknown ssid>", "", "[ssid",
                ) and not ssid.startswith("00:") and not ssid.startswith("<"):
                    if ssid not in ssids:
                        ssids.append(ssid)
                        continue

            # 格式 2: BSSID 后面跟的 SSID（某些 ROM 的表格格式）
            # 如: xx:xx:xx:xx:xx:xx  MyWiFi  ...
            m2 = re.match(
                r'^[0-9a-fA-F]{2}:[0-9a-fA-F]{2}:[0-9a-fA-F]{2}:[0-9a-fA-F]{2}:[0-9a-fA-F]{2}:[0-9a-fA-F]{2}\s+(\S+)',
                line,
            )
            if m2:
                ssid = m2.group(1)
                if ssid and len(ssid) >= 1 and ssid not in ssids and ssid != "*":
                    ssids.append(ssid)
                    continue

            # 格式 3: "SSID" 用双引号括起来
            m3 = re.findall(r'"([^"]{1,32})"', line)
            for s in m3:
                s = s.strip()
                if s and s not in ssids and len(s) < 33:
                    ssids.append(s)

        return ssids

    @staticmethod
    def trigger_scan():
        """触发一次 WiFi 扫描。"""
        AdbHelper.run("cmd wifi start-scan", timeout=5)

    @staticmethod
    def verify_toggle_works() -> bool:
        """验证 WiFi 开关是否正常（能关闭再打开）。"""
        was_on = WifiDetector.is_wifi_enabled()
        if not was_on:
            AdbHelper.run("svc wifi enable", timeout=5)
            time.sleep(3)
        # 关闭
        AdbHelper.run("svc wifi disable", timeout=5)
        time.sleep(3)
        off_ok = not WifiDetector.is_wifi_enabled()
        # 打开
        AdbHelper.run("svc wifi enable", timeout=5)
        time.sleep(3)
        on_ok = WifiDetector.is_wifi_enabled()
        return off_ok and on_ok

    # ---- 综合健康检测 ----

    @classmethod
    def check_health(cls, ssid: Optional[str], password: Optional[str]) -> Tuple[bool, str, int, bool]:
        """
        综合健康检测。
        返回 (healthy, state_desc, scan_count, connect_ok)。

        判定顺序：
        1. WiFi 是否已开启
        2. 能否扫描到 AP（最多重试 3 次）
        3. 如果指定了 ssid，能否连接到目标热点
        """
        print("  [检测] 检查 WiFi 开关状态...")
        if not cls.is_wifi_enabled():
            return False, "disabled", 0, False

        # 扫描 + 重试（开机后初次扫描可能较慢）
        ssids = []
        for attempt in range(1, 4):
            print(f"  [检测] 触发 WiFi 扫描... (第 {attempt}/3 次)")
            cls.trigger_scan()
            time.sleep(SCAN_WAIT)

            ssids = cls.get_scan_results()
            if ssids:
                print(f"  [检测] 扫描到 {len(ssids)} 个 AP")
                break
            if attempt < 3:
                print(f"  [检测] 扫描结果为空，{SCAN_WAIT}s 后重试...")
                time.sleep(SCAN_WAIT)

        if not ssids:
            return False, "no_scan", 0, False

        # 连接验证
        if ssid:
            print(f"  [检测] 尝试连接 {ssid}...")
            ok = ConnectionTester.connect_and_verify(ssid, password)
            if not ok:
                return False, "connect_fail", len(ssids), False
            ConnectionTester.disconnect()
            return True, "healthy", len(ssids), True

        return True, "healthy", len(ssids), False

    # ---- 基线采集 ----

    @staticmethod
    def collect_baseline(ssid: Optional[str] = None) -> dict:
        """采集注入前的基线数据：AP数量、连接状态。"""
        WifiDetector.trigger_scan()
        time.sleep(SCAN_WAIT)
        scan = WifiDetector.get_scan_results()
        connected = False
        if ssid:
            connected = WifiDetector.get_connected_ssid() == ssid
        return {
            "ap_count": len(scan),
            "connected": connected,
            "wifi_enabled": WifiDetector.is_wifi_enabled(),
        }

    @staticmethod
    def get_connected_ssid() -> Optional[str]:
        """获取当前连接的 WiFi SSID。支持多种 ROM 格式。"""
        # 方式1: dumpsys wifi (最通用)
        _, out, _ = AdbHelper.run("dumpsys wifi | grep 'mWifiInfo'", timeout=10)
        if out:
            for pat in [
                r'SSID:\s*"?([^",\n]+)"?',
                r'ssid[=\s]+"?([^",\n]+)"?',
                r'"([^"]+)"',  # 最后兜底：取第一对引号内容
            ]:
                m = re.search(pat, out, re.IGNORECASE)
                if m:
                    val = m.group(1).strip()
                    if val and val not in ("<unknown ssid>", "null", "") and len(val) < 64:
                        return val
        # 方式2: cmd wifi status
        _, out, _ = AdbHelper.run("cmd wifi status", timeout=5)
        if out:
            for line in out.split("\n"):
                if "SSID" in line or "ssid" in line:
                    m = re.search(r'"?([^"\n]{1,32})"?\s*$', line.strip())
                    if m:
                        val = m.group(1).strip()
                        if val and val not in ("<unknown ssid>", "null", ""):
                            return val
        # 方式3: dumpsys netstats (部分 ROM)
        _, out, _ = AdbHelper.run("dumpsys netstats | grep 'iface=wlan'", timeout=5)
        if out:
            m = re.search(r'networkId[= ]+(\S+)', out)
            if m:
                net_id = m.group(1).strip()
                # 用 network ID 查 SSID
                _, out2, _ = AdbHelper.run("cmd wifi list-networks", timeout=5)
                if out2:
                    for line in out2.split("\n"):
                        if net_id in line:
                            m2 = re.search(r'"([^"]+)"', line)
                            if m2:
                                return m2.group(1).strip()
        return None


# ============================================================
# 类 3: FaultInjector — 主动注入故障
# ============================================================

class FaultInjector:
    """模拟 WiFi 异常，支持多种故障类型随机注入。"""

    # 故障类型定义：标识 → (名称, 命令列表, 等待时间, 描述)
    FAULT_TYPES: Dict[str, Tuple[str, List[Tuple[str, int]], int, str]] = {
        "wpa_kill": (
            "杀掉wpa_supplicant",
            [
                ("svc wifi disable", 5),
                ("killall wpa_supplicant 2>/dev/null; pkill -9 wpa_supplicant 2>/dev/null", 5),
            ],
            3,
            "SIGKILL wpa_supplicant + 关闭WiFi",
        ),
        "wpa_freeze": (
            "卡死wpa_supplicant",
            [
                ("kill -19 $(pidof wpa_supplicant) 2>/dev/null", 3),
            ],
            2,
            "SIGSTOP 挂起进程 → 按钮卡死无反应",
        ),
        "wificond_kill": (
            "杀掉wificond",
            [
                ("svc wifi disable", 5),
                ("pkill -9 wificond 2>/dev/null; pkill -9 wificond 2>/dev/null", 5),
            ],
            3,
            "SIGKILL WiFi守护进程",
        ),
    }

    @staticmethod
    def list_types() -> Dict[str, str]:
        """返回故障类型标识 → 名称的映射。"""
        return {k: v[0] for k, v in FaultInjector.FAULT_TYPES.items()}

    @staticmethod
    def inject(fault_id: Optional[str] = None) -> Tuple[bool, str]:
        """
        注入 WiFi 故障。
        如果指定 fault_id 则使用指定类型，否则随机选择。
        返回 (故障是否生效, 故障类型标识)。
        """
        if fault_id and fault_id in FaultInjector.FAULT_TYPES:
            fid = fault_id
        else:
            fid = random.choice(list(FaultInjector.FAULT_TYPES.keys()))

        name, commands, wait_time, desc = FaultInjector.FAULT_TYPES[fid]
        print(f"    [注入故障] 类型: {name}")
        print(f"    [注入故障] 说明: {desc}")

        # 执行所有注入命令
        for cmd, timeout in commands:
            print(f"    [注入故障] 执行: {cmd}")
            AdbHelper.run(cmd, timeout=timeout)
            time.sleep(1)

        time.sleep(wait_time)

        # 验证故障生效
        wifi_on = WifiDetector.is_wifi_enabled()
        if wifi_on:
            # 再尝试关一次
            AdbHelper.run("cmd wifi set-wifi-enabled disabled", timeout=5)
            time.sleep(2)
            wifi_on = WifiDetector.is_wifi_enabled()

        scan = WifiDetector.get_scan_results()

        # 判定：WiFi关掉或扫描为空
        if fid == "wpa_freeze":
            fault_effective = not wifi_on or len(scan) == 0
        else:
            fault_effective = not wifi_on or len(scan) == 0

        if fault_effective:
            state = "disabled" if not wifi_on else "scan_empty"
            print(f"    [注入故障] 故障已生效 (WiFi={'关' if not wifi_on else '开'}, AP={len(scan)})")
        else:
            print(f"    [注入故障] WARN: 故障可能未完全生效 (WiFi={'开' if wifi_on else '关'}, AP={len(scan)})")

        return fault_effective, fid


# ============================================================
# 类 4: RecoveryMethods — 两种恢复方案
# ============================================================

class RecoveryMethods:
    """方案一：开关 WiFi；方案二：飞行模式切换。"""

    _settings_intent = DEFAULT_SETTINGS_INTENT
    """uiautomator2 降级时打开的设置页面，可通过 configure() 覆盖"""

    @classmethod
    def configure(cls, settings_intent: str):
        cls._settings_intent = settings_intent

    # ---- 方案一：开关 WiFi ----

    @classmethod
    def _wifi_off(cls) -> bool:
        """关闭 WiFi，多命令尝试，返回是否成功关闭。"""
        for cmd, desc in [
            ("svc wifi disable", "svc wifi disable"),
            ("cmd wifi set-wifi-enabled disabled", "cmd wifi"),
        ]:
            rc, _, _ = AdbHelper.run(cmd, timeout=5)
            time.sleep(2)
            if not WifiDetector.is_wifi_enabled():
                global _wifi_cmd_mode
                _wifi_cmd_mode = "svc" if "svc" in cmd else "cmd_wifi"
                print(f"    [关闭WiFi] {desc} → 已关闭")
                return True
        return False

    @classmethod
    def _wifi_on(cls) -> bool:
        """打开 WiFi，多命令尝试，返回是否成功开启。"""
        for cmd, desc in [
            ("svc wifi enable", "svc wifi enable"),
            ("cmd wifi set-wifi-enabled enabled", "cmd wifi"),
        ]:
            rc, _, _ = AdbHelper.run(cmd, timeout=5)
            time.sleep(2)
            if WifiDetector.is_wifi_enabled():
                print(f"    [开启WiFi] {desc} → 已开启")
                return True
        # 全部失败，降级 uiautomator2
        cls._toggle_wifi_via_ui()
        time.sleep(2)
        return WifiDetector.is_wifi_enabled()

    @classmethod
    def toggle_wifi(cls) -> Tuple[bool, int]:
        """
        关闭 WiFi → 等待 3s → 打开 WiFi → 等待 5s → 验证扫描。
        返回 (success, elapsed_ms)。
        """
        start = time.time()

        # ① 关闭 WiFi（验证状态确实变了）
        print("    [方案一] 关闭 WiFi...")
        if not cls._wifi_off():
            print("    [方案一] WARN: 无法关闭 WiFi，尝试继续...")
        time.sleep(WIFI_TOGGLE_OFF_WAIT)

        # ② 打开 WiFi
        print("    [方案一] 重新打开 WiFi...")
        if not cls._wifi_on():
            print("    [方案一] FAIL: 无法打开 WiFi")
            elapsed = int((time.time() - start) * 1000)
            return False, elapsed
        time.sleep(WIFI_TOGGLE_ON_WAIT)

        # ③ 验证恢复（扫描 + 开关状态）
        wifi_on = WifiDetector.is_wifi_enabled()
        scan = []
        if wifi_on:
            WifiDetector.trigger_scan()
            time.sleep(SCAN_WAIT)
            scan = WifiDetector.get_scan_results()
            print(f"    [方案一] 恢复后扫描到 {len(scan)} 个 AP")
        else:
            print("    [方案一] WARN: WiFi 开关状态仍为关闭")

        elapsed = int((time.time() - start) * 1000)
        return wifi_on and len(scan) > 0, elapsed

    @classmethod
    def _toggle_wifi_via_ui(cls):
        """uiautomator2 降级：打开设置页面点击 WiFi 开关。"""
        try:
            import uiautomator2 as u2
            d = u2.connect()
            d.shell(f"am start -a {cls._settings_intent}")
            time.sleep(2)

            switch = None
            for rid in [
                "android:id/switch_widget",
                "com.android.settings:id/switch_widget",
                "android:id/switchWidget",
                "com.android.settings:id/switch_bar",
            ]:
                el = d(resourceId=rid)
                if el.exists:
                    switch = el
                    break
            if switch is None:
                el = d(className="android.widget.Switch")
                if el.exists:
                    switch = el

            if switch is not None:
                switch.click()
                time.sleep(1)
                switch.click()
                time.sleep(2)

            d.press("home")
        except ImportError:
            pass
        except Exception:
            pass

    # ---- 方案二：飞行模式切换 ----

    @classmethod
    def toggle_airplane_mode(cls) -> Tuple[bool, int]:
        """
        开启飞行模式 → 等待 3s → 关闭飞行模式 → 等待 8s → 验证。
        返回 (success, elapsed_ms)。
        """
        start = time.time()

        # ① 开启
        AdbHelper.run("settings put global airplane_mode_on 1", timeout=5)
        AdbHelper.run(
            "am broadcast -a android.intent.action.AIRPLANE_MODE --ez state true",
            timeout=5,
        )
        time.sleep(AIRPLANE_ON_WAIT)

        # ② 关闭
        AdbHelper.run("settings put global airplane_mode_on 0", timeout=5)
        AdbHelper.run(
            "am broadcast -a android.intent.action.AIRPLANE_MODE --ez state false",
            timeout=5,
        )
        time.sleep(AIRPLANE_OFF_WAIT)

        # ③ 确保 WiFi 开启并扫描
        if not WifiDetector.is_wifi_enabled():
            AdbHelper.run("svc wifi enable", timeout=5)
            time.sleep(SCAN_WAIT)

        WifiDetector.trigger_scan()
        time.sleep(SCAN_WAIT)
        scan = WifiDetector.get_scan_results()

        elapsed = int((time.time() - start) * 1000)
        return len(scan) > 0, elapsed


# ============================================================
# 类 4: ConnectionTester — 连接验证
# ============================================================

class ConnectionTester:
    """验证是否能连接预先保存的热点。"""

    @staticmethod
    def is_saved(ssid: str) -> bool:
        _, out, _ = AdbHelper.run(f"cmd wifi list-networks | grep '{ssid}'", timeout=10)
        return ssid in out

    @staticmethod
    def connect_and_verify(ssid: str, password: Optional[str] = None, timeout: int = 15) -> bool:
        """连接到目标 WiFi 并验证是否成功。

        如果有密码则主动连接；如果无密码但网络已保存，则依赖 Android 自动回连。
        """
        if password:
            # 主动连接
            rc, out, err = AdbHelper.run(
                f'cmd wifi connect-network "{ssid}" wpa2 "{password}"',
                timeout=timeout,
            )
            if rc != 0 and "error" in (out + err).lower():
                return False
            time.sleep(CONNECT_WAIT)
        else:
            # 无密码，依赖自动回连（已保存网络）
            # Android 开启 WiFi 后会自动连接已保存的网络，等待一下
            print(f"    [连接] 无密码，等待自动回连已保存网络...")
            time.sleep(CONNECT_WAIT + 3)

        # 方式 1：dumpsys
        _, out, _ = AdbHelper.run("dumpsys wifi | grep 'mWifiInfo'", timeout=10)
        if ssid in out.replace('"', ""):
            return True

        # 方式 2：cmd wifi status
        _, out, _ = AdbHelper.run("cmd wifi status", timeout=5)
        if ssid in out:
            return True

        return False

    @staticmethod
    def disconnect():
        AdbHelper.run("cmd wifi disconnect", timeout=5)

    @staticmethod
    def get_saved_password(ssid: str) -> Optional[str]:
        """从设备读取已保存 WiFi 密码。尝试多条路径，兼容不同 ROM。"""
        # 路径列表（按优先级）
        paths = [
            # Android 标准路径
            "/data/misc/wifi/wpa_supplicant.conf",
            # 部分厂商 ROM
            "/data/vendor/wifi/wpa/wpa_supplicant.conf",
            "/data/misc/wifi/WifiConfigStore.xml",
            # Android 11+ Apex
            "/data/misc/apexdata/com.android.wifi/WifiConfigStore.xml",
        ]

        for path in paths:
            # 先尝试直接读取
            rc, out, _ = AdbHelper.run(
                f"cat {path} 2>/dev/null", timeout=10
            )
            if not out or rc != 0:
                # 尝试 su 提权
                rc, out, _ = AdbHelper.run(
                    f"su -c 'cat {path}' 2>/dev/null", timeout=10
                )
            if not out:
                continue

            # 解析 wpa_supplicant.conf 格式
            pwd = cls._parse_wpa_supplicant(out, ssid)
            if pwd:
                return pwd

            # 解析 WifiConfigStore.xml 格式
            pwd = cls._parse_wifi_config_xml(out, ssid)
            if pwd:
                return pwd

        return None

    @staticmethod
    def _parse_wpa_supplicant(text: str, target_ssid: str) -> Optional[str]:
        """解析 wpa_supplicant.conf 中的密码。"""
        # 按 network={ 分块
        blocks = re.split(r'\n(?=network=)', text)
        for block in blocks:
            # 匹配 ssid
            m_ssid = re.search(r'ssid\s*=\s*"?([^"\n]+)"?', block)
            if not m_ssid:
                continue
            # SSID 可能以 hex 编码: 4d7957694669
            block_ssid = m_ssid.group(1).strip()
            if block_ssid != target_ssid:
                # 尝试解码 hex 编码的 SSID
                try:
                    decoded = codecs.decode(block_ssid, 'hex').decode('utf-8', errors='replace')
                    if decoded != target_ssid:
                        continue
                except Exception:
                    continue
            # psk (WPA/WPA2)
            m_psk = re.search(r'psk\s*=\s*"?([^"\n]+)"?', block)
            if m_psk:
                psk = m_psk.group(1).strip()
                # 过滤掉占位符
                if psk and psk not in ("password", "changeme", "null"):
                    return psk
            # sae_password (WPA3)
            m_sae = re.search(r'sae_password\s*=\s*"?([^"\n]+)"?', block)
            if m_sae:
                return m_sae.group(1).strip()
            # wep_key0 (WEP)
            m_wep = re.search(r'wep_key0\s*=\s*"?([^"\n]+)"?', block)
            if m_wep:
                return m_wep.group(1).strip()
        return None

    @staticmethod
    def _parse_wifi_config_xml(text: str, target_ssid: str) -> Optional[str]:
        """解析 WifiConfigStore.xml 中的密码（Android 10+）。"""
        # 查找 <string name="SSID">xxx</string> 后面的 PreSharedKey
        escaped_ssid = target_ssid.replace('"', '&quot;')
        pattern = rf'<string\s+name="SSID">"?{re.escape(target_ssid)}"?</string>'
        m = re.search(pattern, text)
        if not m:
            # 尝试 hex 编码
            hex_ssid = codecs.encode(target_ssid.encode('utf-8'), 'hex').decode('ascii')
            pattern2 = rf'<string\s+name="SSID">"?{re.escape(hex_ssid)}"?</string>'
            m = re.search(pattern2, text)
        if m:
            # 在后面找 PreSharedKey
            after = text[m.end():m.end() + 500]
            m_psk = re.search(r'<string\s+name="PreSharedKey">"?([^"<]+)"?</string>', after)
            if m_psk:
                return m_psk.group(1).strip()
        return None


# ============================================================
# 类 5: TestRunner — 测试编排
# ============================================================

class TestRunner:
    """按配置执行多轮重启+WiFi 自愈测试。"""

    def __init__(
        self,
        cycles: int = DEFAULT_CYCLES,
        boot_wait: int = DEFAULT_BOOT_WAIT,
        ssid: Optional[str] = None,
        password: Optional[str] = None,
        auto_detect: bool = True,
        output_dir: str = DEFAULT_OUTPUT,
        settings_intent: str = DEFAULT_SETTINGS_INTENT,
        recovery_mode: str = DEFAULT_RECOVERY_MODE,
    ):
        self.cycles = cycles
        self.boot_wait = boot_wait
        self.ssid = ssid
        self.password = password
        self.auto_detect = auto_detect
        self.output_dir = output_dir
        self.records: List[CycleRecord] = []
        self.device_info = ""
        self.start_time = ""
        self.recovery_mode = recovery_mode if recovery_mode in RECOVERY_MODES else DEFAULT_RECOVERY_MODE

        RecoveryMethods.configure(settings_intent)

    def run(self):
        """主测试入口。"""
        print("正在检查 ADB 连接...")
        if not AdbHelper.check_device():
            print("[ERROR] 未检测到已连接的设备，请确认：")
            print("  1. 设备已通过 USB 连接")
            print("  2. 已开启 USB 调试")
            print("  3. adb devices 可见且状态为 device")
            sys.exit(1)

        self.device_info = AdbHelper.get_device_info()
        self.start_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # 自动检测WiFi凭据
        if self.auto_detect and not self.ssid:
            print("自动检测当前 WiFi 连接...")
            detected = WifiDetector.get_connected_ssid()
            if detected:
                self.ssid = detected
                print(f"  检测到 SSID: {self.ssid}")
                if not self.password:
                    pwd = ConnectionTester.get_saved_password(self.ssid)
                    if pwd:
                        self.password = pwd
                        print(f"  自动获取密码成功")
                    else:
                        print(f"  (无法自动读取密码，将依赖自动回连)")
            else:
                print(f"  未检测到当前连接的 WiFi，连接验证将跳过")

        print(f"设备: {self.device_info}")
        print(f"开始时间: {self.start_time}")
        print(f"循环次数: {self.cycles}  |  启动等待: {self.boot_wait}s")
        if self.ssid:
            print(f"连接验证: {self.ssid}")
            saved = ConnectionTester.is_saved(self.ssid)
            if not saved:
                print(f"  [WARN] '{self.ssid}' 未在已保存网络中找到，连接验证将跳过")
            else:
                has_pwd = bool(self.password)
                if has_pwd:
                    print(f"  密码: 已获取 → 主动连接验证")
                else:
                    print(f"  密码: 无 → 依赖自动回连验证（已保存网络）")
        print("=" * 60)

        for i in range(1, self.cycles + 1):
            record = self._run_cycle(i)
            self.records.append(record)
            self._print_cycle_result(record, self.cycles)

        return self.records

    # ------------------------------------------------------------
    # 单轮测试：降级链流程
    # ------------------------------------------------------------

    def _run_cycle(self, cycle_num: int) -> CycleRecord:
        """
        单轮测试流程（根据 recovery_mode 切换策略）：
        ① 重启设备 + 等待启动完成
        ② 采集基线（AP数、连接状态）
        ③ 注入故障（随机类型）
        ④ 方案一/方案二 根据模式执行:
           - degrade: 方案一 → 失败则方案二（降级链）
           - both:    方案一 → 重新注入同类型故障 → 方案二（全测）
        """
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        record = CycleRecord(cycle=cycle_num, timestamp=ts)

        # 创建本轮日志目录
        log_dir = os.path.join(self.output_dir, "logs", f"run_{self.start_time.replace(':', '').replace(' ', '_')}", f"iteration_{cycle_num}")
        os.makedirs(log_dir, exist_ok=True)
        record.log_path = log_dir

        print(f"\n{'─' * 40}")
        print(f"[{cycle_num}/{self.cycles}] 正在重启设备...")

        # ① 重启 + 等待启动
        reboot_start = time.time()
        record.reboot_ok = AdbHelper.reboot()
        if not record.reboot_ok:
            record.m1_detail = "重启失败"; record.m2_detail = "重启失败"
            return record
        if not AdbHelper.wait_for_device():
            record.m1_detail = "启动超时"; record.m2_detail = "启动超时"
            return record
        if not AdbHelper.wait_for_boot_completed():
            record.m1_detail = "boot_completed超时"; record.m2_detail = "boot_completed超时"
            return record

        boot_time = datetime.now().strftime("%H:%M:%S")
        record.boot_completed_time = boot_time
        print(f"  Android开机完成: {boot_time}")

        print(f"  等待 {self.boot_wait}s 系统初始化...")
        for s in range(self.boot_wait, 0, -5):
            sys.stdout.write(f"\r  剩余 {s}s ...")
            sys.stdout.flush()
            time.sleep(min(5, s))
        print()

        wifi_ready_start = time.time()
        WifiDetector.wait_for_wifi_ready()
        wifi_ready_time = datetime.now().strftime("%H:%M:%S")
        record.wifi_ready_time = wifi_ready_time
        record.wifi_init_duration_s = time.time() - wifi_ready_start + self.boot_wait
        print(f"  WiFi列表加载完成: {wifi_ready_time} (初始化耗时 {record.wifi_init_duration_s:.1f}s)")

        # ② 采集基线
        print(f"  采集基线数据...")
        baseline = WifiDetector.collect_baseline(self.ssid)
        record.baseline_ap = baseline["ap_count"]
        record.baseline_connected = baseline["connected"]
        print(f"  基线: AP={record.baseline_ap}, "
              f"WiFi={'开' if baseline['wifi_enabled'] else '关'}"
              + (f", 已连接={record.baseline_connected}" if self.ssid else ""))

        # ③ 注入故障
        print(f"  {'─' * 30}")
        record.fault_ok, record.fault_type = FaultInjector.inject()
        fault_name = FaultInjector.FAULT_TYPES.get(record.fault_type, ('?',))[0]
        print(f"  故障类型: {fault_name}  |  恢复模式: {RECOVERY_MODES[self.recovery_mode]}")

        mode = self.recovery_mode

        # ④ 方案一：始终执行
        print(f"  >>> 方案一：开关 WiFi <<<")
        self._verify_recovery(record, method=1)

        # ⑤ 方案二
        if mode == "both":
            # 无论方案一是否成功，都再测方案二（重新注入同一故障类型）
            print(f"  >>> 方案二：飞行模式（全测模式） <<<")
            if record.fault_ok:
                print(f"    重新注入同类型故障...")
                FaultInjector.inject(record.fault_type)
            record.m2_skipped = False
            self._verify_recovery(record, method=2)
        else:
            # degrade 模式：方案一失败才走方案二
            if not record.m1_recovered:
                print(f"  >>> 方案二：飞行模式（降级） <<<")
                record.m2_skipped = False
                self._verify_recovery(record, method=2)
            else:
                print(f"  方案一成功，跳过方案二")
                record.m2_skipped = True

        # ⑥ 最终状态汇总
        record.scan_empty = record.final_scan_count == 0

        # 保存本轮日志
        record_data = record.to_dict()
        with open(os.path.join(log_dir, "record.json"), "w", encoding="utf-8") as f:
            json.dump(record_data, f, ensure_ascii=False, indent=2)

        return record

    # ------------------------------------------------------------
    # 恢复 + 验证（方案一和方案二共用）
    # ------------------------------------------------------------

    def _verify_recovery(self, record: CycleRecord, method: int):
        """执行恢复方案并验证结果。method=1 方案一, method=2 方案二。"""
        prefix = "方案一" if method == 1 else "方案二"

        # 执行恢复
        print(f"  [{prefix}] 执行恢复...")
        if method == 1:
            ok, elapsed = RecoveryMethods.toggle_wifi()
        else:
            ok, elapsed = RecoveryMethods.toggle_airplane_mode()

        if method == 1:
            record.m1_time_ms = elapsed
        else:
            record.m2_time_ms = elapsed

        if not ok:
            detail = f"{prefix}恢复失败 (无法完成操作)"
            if method == 1:
                record.m1_detail = detail
                record.m1_recovered = False
            else:
                record.m2_detail = detail
                record.m2_recovered = False
            return

        # 验证恢复后状态
        print(f"  [{prefix}] 验证恢复后状态...")
        toggle_ok = WifiDetector.verify_toggle_works()
        print(f"    [验证] 开关可点击: {'OK' if toggle_ok else 'FAIL'}")

        # 扫描验证（带重试）
        scan = []
        for retry in range(1, 4):
            WifiDetector.trigger_scan()
            time.sleep(SCAN_WAIT)
            scan = WifiDetector.get_scan_results()
            if scan:
                break
            if retry < 3:
                print(f"    [验证] 扫描为空，第{retry}次重试...")
                time.sleep(3)

        scan_count = len(scan)
        print(f"    [验证] AP 列表: {scan_count} 个 (基线: {record.baseline_ap})")

        # 恢复判定：AP >= 基线80%
        min_ap = max(1, int(record.baseline_ap * 0.8))
        recovered = scan_count >= min_ap

        # 连接验证
        connect_ok = True
        if self.ssid and scan:
            # 先确认网络是否已保存
            is_saved = ConnectionTester.is_saved(self.ssid)
            if self.password or is_saved:
                connect_ok = ConnectionTester.connect_and_verify(self.ssid, self.password)
                if connect_ok:
                    ConnectionTester.disconnect()
                print(f"    [验证] 连接 {self.ssid}: {'OK' if connect_ok else 'FAIL'}")
            else:
                print(f"    [验证] {self.ssid} 未保存且无密码，跳过连接验证")
                connect_ok = True  # 非关键失败，不标记为 FAIL
        elif not self.ssid:
            connect_ok = True  # 无SSID，跳过

        detail = (
            f"[{FaultInjector.FAULT_TYPES.get(record.fault_type, ('?',))[0]}] "
            f"基线AP={record.baseline_ap}, 恢复AP={scan_count}, "
            f"开关={'OK' if toggle_ok else 'FAIL'}"
            + (f", 连接={'OK' if connect_ok else 'FAIL'}" if self.ssid and self.password else "")
        )

        if method == 1:
            record.m1_recovered = recovered
            record.m1_toggle_ok = toggle_ok
            record.m1_scan_after = scan_count
            record.m1_connect_ok = connect_ok
            record.m1_detail = detail
        else:
            record.m2_recovered = recovered
            record.m2_toggle_ok = toggle_ok
            record.m2_scan_after = scan_count
            record.m2_connect_ok = connect_ok
            record.m2_detail = detail

    # ------------------------------------------------------------
    # 终端输出
    # ------------------------------------------------------------

    def _print_cycle_result(self, record: CycleRecord, total: int):
        m1_icon = "OK" if record.m1_recovered else "FAIL"
        ft = FaultInjector.FAULT_TYPES.get(record.fault_type, ('?',))[0]

        if not record.m2_skipped:
            m2_icon = "OK" if record.m2_recovered else "FAIL"
            m2_info = f"方案二:{m2_icon:4s} 开关:{'OK' if record.m2_toggle_ok else 'FAIL':4s} AP:{record.m2_scan_after:3d}/{record.baseline_ap}"
        else:
            m2_info = "方案二:SKIP"

        conn1 = f" CONN:{'OK' if record.m1_connect_ok else 'FAIL'}" if (self.ssid and self.password) else ""

        print(
            f"[{record.cycle:3d}/{total}] "
            f"故障:{ft:14s} 基线AP:{record.baseline_ap:3d} | "
            f"方案一:{m1_icon:4s} 开关:{'OK' if record.m1_toggle_ok else 'FAIL':4s} "
            f"AP:{record.m1_scan_after:3d}/{record.baseline_ap}{conn1} | "
            f"{m2_info}"
        )


# ============================================================
# 类 6: ReportGenerator — 报告生成
# ============================================================

class ReportGenerator:
    """终端打印 + JSON + HTML 三份报告。"""

    def __init__(
        self, records: List[CycleRecord],
        device_info: str, start_time: str, output_dir: str,
        ssid: Optional[str] = None,
        recovery_mode: str = DEFAULT_RECOVERY_MODE,
    ):
        self.records = records
        self.device_info = device_info
        self.start_time = start_time
        self.output_dir = output_dir
        self.ssid = ssid
        self.recovery_mode = recovery_mode

    def generate_all(self):
        os.makedirs(self.output_dir, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.save_json(os.path.join(self.output_dir, f"wifi_healing_{ts}.json"))
        self.save_html(os.path.join(self.output_dir, f"wifi_healing_{ts}.html"))
        self.save_excel(os.path.join(self.output_dir, f"wifi_healing_{ts}.xlsx"))
        self.print_terminal()
        print(f"\n报告已保存到: {self.output_dir}")

    def _compute_stats(self) -> dict:
        total = len(self.records)
        m1_recovered = sum(1 for r in self.records if r.m1_recovered)
        m1_toggle_ok = sum(1 for r in self.records if r.m1_toggle_ok)
        m1_connect_ok = sum(1 for r in self.records if r.m1_connect_ok)
        m1_avg = sum(r.m1_time_ms for r in self.records if r.m1_recovered) / max(m1_recovered, 1)
        m1_avg_pct = sum(r.m1_recovery_pct for r in self.records) / total if total > 0 else 0

        m2_attempted = sum(1 for r in self.records if not r.m2_skipped)
        m2_recovered = sum(1 for r in self.records if r.m2_recovered)
        m2_toggle_ok = sum(1 for r in self.records if r.m2_toggle_ok)
        m2_connect_ok = sum(1 for r in self.records if r.m2_connect_ok)
        m2_avg = sum(r.m2_time_ms for r in self.records if r.m2_recovered) / max(m2_recovered, 1)
        m2_avg_pct = sum(r.m2_recovery_pct for r in self.records if not r.m2_skipped) / max(m2_attempted, 1)

        # 降级链统计
        degraded = sum(1 for r in self.records if not r.m2_skipped)  # m1失败→走m2
        degraded_rescued = sum(1 for r in self.records if not r.m2_skipped and r.m2_recovered)  # m2救回
        both_fail = sum(1 for r in self.records if not r.m1_recovered and not r.m2_recovered)  # 双双失败

        # 按故障类型统计（只看方案一）
        fault_stats: Dict[str, Dict] = {}
        for r in self.records:
            ft = r.fault_type
            if ft not in fault_stats:
                fault_stats[ft] = {"name": FaultInjector.FAULT_TYPES.get(ft, ('?',))[0], "total": 0, "m1_ok": 0, "m2_rescued": 0}
            fault_stats[ft]["total"] += 1
            if r.m1_recovered:
                fault_stats[ft]["m1_ok"] += 1
            if not r.m2_skipped and r.m2_recovered:
                fault_stats[ft]["m2_rescued"] += 1
        for ft in fault_stats:
            t = fault_stats[ft]["total"]
            fs = fault_stats[ft]
            fs["rate"] = (fs["m1_ok"] + fs["m2_rescued"]) / t * 100 if t > 0 else 0
            fs["m1_rate"] = fs["m1_ok"] / t * 100 if t > 0 else 0

        overall = (m1_recovered + degraded_rescued) / total * 100 if total > 0 else 0

        # 其他统计
        reboot_ok_count = sum(1 for r in self.records if r.reboot_ok)
        connect_ok_count = sum(1 for r in self.records if r.final_connect_ok)

        return {
            "total": total,
            "m1_recovered": m1_recovered, "m1_toggle_ok": m1_toggle_ok,
            "m1_connect_ok": m1_connect_ok,
            "m1_rate": m1_recovered / total * 100 if total > 0 else 0,
            "m1_avg_ms": m1_avg, "m1_avg_pct": m1_avg_pct,
            "m2_attempted": m2_attempted, "m2_recovered": m2_recovered,
            "m2_toggle_ok": m2_toggle_ok, "m2_connect_ok": m2_connect_ok,
            "m2_rate": m2_recovered / m2_attempted * 100 if m2_attempted > 0 else 0,
            "m2_avg_ms": m2_avg, "m2_avg_pct": m2_avg_pct,
            "degraded": degraded, "degraded_rescued": degraded_rescued,
            "degraded_rescue_rate": degraded_rescued / degraded * 100 if degraded > 0 else 0,
            "both_fail": both_fail,
            "overall_rate": overall,
            "reboot_ok": reboot_ok_count,
            "connect_ok": connect_ok_count,
            "fault_stats": fault_stats,
        }

    def print_terminal(self):
        s = self._compute_stats()
        print()
        print("=" * 80)
        print("                  WiFi 异常自愈能力测试报告")
        print("=" * 80)
        print(f"  测试时间: {self.start_time}")
        print(f"  设备:     {self.device_info}")
        print(f"  总循环:   {s['total']}")
        print(f"  恢复策略: {RECOVERY_MODES.get(self.recovery_mode, self.recovery_mode)}")
        if self.ssid:
            print(f"  目标SSID: {self.ssid}")
        print("-" * 80)
        # 方案一
        print(f"  方案一（开关WiFi）:     恢复 {s['m1_recovered']}/{s['total']} "
              f"({s['m1_rate']:.1f}%)  |  平均耗时 {s['m1_avg_ms'] / 1000:.1f}s  |  平均恢复率 {s['m1_avg_pct']:.0f}%")
        # 方案二
        m2_label = "方案二（飞行模式/全测）" if self.recovery_mode == "both" else "方案二（飞行模式/降级）"
        print(f"  {m2_label}: 执行 {s['m2_attempted']} 次  |  恢复 {s['m2_recovered']}/{s['m2_attempted']} "
              f"({s['m2_rate']:.1f}%)  |  平均耗时 {s['m2_avg_ms'] / 1000:.1f}s")
        print("-" * 80)

        # 策略统计
        if self.recovery_mode == "degrade":
            print(f"  降级链: 方案一失败 {s['degraded']} 次 → 方案二救回 {s['degraded_rescued']} 次 "
                  f"({s['degraded_rescue_rate']:.1f}%)  |  双双失败: {s['both_fail']}")
        else:
            both_m1_ok = sum(1 for r in self.records if r.m1_recovered)
            both_m2_ok = sum(1 for r in self.records if r.m2_recovered)
            both_ok = sum(1 for r in self.records if r.m1_recovered and r.m2_recovered)
            print(f"  全测模式: M1成功={both_m1_ok}  |  M2成功={both_m2_ok}  |  两者都成功={both_ok}  |  双双失败={s['both_fail']}")
        print("-" * 80)

        # 按故障类型统计
        if s.get("fault_stats"):
            print(f"  按故障类型:")
            print(f"  {'故障类型':20s} {'出现':5s} {'方案一OK':10s} {'方案二救回':10s} {'综合成功率':10s}")
            for ft, fs in sorted(s["fault_stats"].items()):
                print(f"  {fs['name']:20s} {fs['total']:>4d}  {fs['m1_ok']:>6d}/{fs['total']:<3d}  "
                      f"{fs['m2_rescued']:>6d}      {fs['rate']:>6.1f}%")
            print("-" * 80)

        rate = s['overall_rate']
        if rate >= 95:
            print(f"  结论: [PASS] 通过 (综合成功率 {rate:.1f}% >= 95%, 无永久性失效)")
        else:
            print(f"  结论: [FAIL] 不通过 (综合成功率 {rate:.1f}% < 95%)")
        print("=" * 80)

        # ---- 逐轮明细表（用户指定格式） ----
        print()
        print("逐轮明细:")
        print("=" * 180)
        hdr = (f"  {'轮次':4s} | {'开始时间':19s} | {'开机完成':8s} | {'WiFi就绪':8s} | "
               f"{'初始化耗时':8s} | {'列表为空':5s} | {'连接成功':5s} | {'本轮成功':5s} | "
               f"{'失败类型':12s} | {'失败详情':20s} | {'重启OK':5s} | {'日志路径'}")
        print(hdr)
        print("-" * 180)
        for r in self.records:
            scan_empty = "Y" if r.scan_empty else "N"
            final_conn = "OK" if r.final_connect_ok else ("-" if not self.ssid else "FAIL")
            success = "OK" if r.success else "FAIL"
            reboot_ok = "OK" if r.reboot_ok else "FAIL"
            fail_type = r.failure_type or "-"
            fail_detail = (r.failure_detail or "-")[:20]
            log_path = r.log_path or "-"
            print(
                f"  {r.cycle:>4d} | {r.timestamp:19s} | {r.boot_completed_time:8s} | {r.wifi_ready_time:8s} | "
                f"{r.wifi_init_duration_s:>6.1f}s   | {scan_empty:>5s} | {final_conn:>5s} | {success:>5s} | "
                f"{fail_type:12s} | {fail_detail:20s} | {reboot_ok:>5s} | {log_path}"
            )
        print("-" * 180)

    def save_json(self, path: str):
        s = self._compute_stats()
        data = {
            "meta": {
                "test_time": self.start_time,
                "device": self.device_info,
                "total_cycles": len(self.records),
                "target_ssid": self.ssid,
                "recovery_mode": self.recovery_mode,
                "fault_types": {k: v[0] for k, v in FaultInjector.FAULT_TYPES.items()},
            },
            "summary": {
                "total_cycles": s["total"],
                "method1_wifi_toggle": {
                    "recovered": s["m1_recovered"], "total": s["total"],
                    "rate": round(s["m1_rate"], 1), "avg_time_ms": round(s["m1_avg_ms"]),
                    "avg_recovery_pct": round(s["m1_avg_pct"], 1),
                    "toggle_ok": s["m1_toggle_ok"], "connect_ok": s["m1_connect_ok"],
                },
                "method2_airplane_mode": {
                    "attempted": s["m2_attempted"],
                    "recovered": s["m2_recovered"],
                    "rate": round(s["m2_rate"], 1), "avg_time_ms": round(s["m2_avg_ms"]),
                    "avg_recovery_pct": round(s["m2_avg_pct"], 1),
                    "toggle_ok": s["m2_toggle_ok"], "connect_ok": s["m2_connect_ok"],
                },
                "degradation_chain": {
                    "m1_fail_count": s["degraded"],
                    "m2_rescued": s["degraded_rescued"],
                    "rescue_rate": round(s["degraded_rescue_rate"], 1),
                    "both_fail": s["both_fail"],
                },
                "overall_rate": round(s["overall_rate"], 1),
                "reboot_ok": s["reboot_ok"],
                "final_connect_ok": s["connect_ok"],
                "fault_stats": {
                    ft: {"name": fs["name"], "total": fs["total"],
                         "m1_ok": fs["m1_ok"], "m2_rescued": fs["m2_rescued"],
                         "rate": round(fs["rate"], 1)}
                    for ft, fs in s.get("fault_stats", {}).items()
                },
            },
            "records": [r.to_dict() for r in self.records],
        }
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    def save_html(self, path: str):
        s = self._compute_stats()
        rate = s['overall_rate']

        # 逐轮明细行
        rows = ""
        for r in self.records:
            cls = "row-fail" if not r.success else ""
            scan_empty = "Y" if r.scan_empty else "N"
            final_conn = "OK" if r.final_connect_ok else ("-" if not self.ssid else "FAIL")
            success = "OK" if r.success else "FAIL"
            reboot_ok = "OK" if r.reboot_ok else "FAIL"
            fail_type = r.failure_type or "-"
            fail_detail = r.failure_detail or "-"
            log_path = r.log_path or "-"
            rows += (
                f"<tr class='{cls}'>"
                f"<td>{r.cycle}</td><td>{r.timestamp[:19]}</td>"
                f"<td>{r.boot_completed_time}</td><td>{r.wifi_ready_time}</td>"
                f"<td>{r.wifi_init_duration_s:.1f}s</td><td>{scan_empty}</td>"
                f"<td>{final_conn}</td><td>{success}</td>"
                f"<td>{fail_type}</td><td>{fail_detail}</td>"
                f"<td>{reboot_ok}</td><td class='log-path'>{log_path}</td></tr>\n"
            )

        # 故障类型统计表格
        fault_rows = ""
        if s.get("fault_stats"):
            for ft, fs in sorted(s["fault_stats"].items()):
                fault_rows += (
                    f"<tr><td>{fs['name']}</td><td>{fs['total']}</td>"
                    f"<td>{fs['m1_ok']}</td><td>{fs['m2_rescued']}</td>"
                    f"<td>{fs['rate']:.1f}%</td></tr>"
                )

        conclusion = (
            f"<span class='pass'>PASS (综合成功率 {rate:.1f}% >= 95%)</span>"
            if rate >= 95 else
            f"<span class='fail'>FAIL (综合成功率 {rate:.1f}% < 95%)</span>"
        )

        html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>WiFi 异常自愈能力测试报告</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:-apple-system,'Segoe UI',Roboto,sans-serif;background:#f5f5f5;color:#333;padding:24px}}
.container{{max-width:1200px;margin:0 auto}}
h1{{font-size:22px;margin-bottom:8px}}
.meta{{color:#888;font-size:13px;margin-bottom:24px}}
.cards{{display:flex;gap:12px;flex-wrap:wrap;margin-bottom:24px}}
.card{{background:#fff;border-radius:8px;padding:16px 20px;flex:1;min-width:140px;box-shadow:0 1px 3px rgba(0,0,0,.08)}}
.card .num{{font-size:28px;font-weight:700}}
.card .label{{font-size:12px;color:#888;margin-top:4px}}
.card.green{{border-left:4px solid #22c55e}}
.card.red{{border-left:4px solid #ef4444}}
.card.blue{{border-left:4px solid #3b82f6}}
.card.orange{{border-left:4px solid #f59e0b}}
.green .num{{color:#22c55e}}
.red .num{{color:#ef4444}}
.blue .num{{color:#3b82f6}}
.orange .num{{color:#f59e0b}}
.section{{background:#fff;border-radius:8px;padding:20px;margin-bottom:16px;box-shadow:0 1px 3px rgba(0,0,0,.08)}}
.section h2{{font-size:16px;margin-bottom:12px;border-bottom:1px solid #eee;padding-bottom:8px}}
table{{width:100%;border-collapse:collapse;font-size:13px}}
th,td{{padding:8px 10px;text-align:left;border-bottom:1px solid #f0f0f0}}
th{{background:#fafafa;font-weight:600;color:#555}}
.row-fail{{background:#fff5f5}}
.pass{{color:#22c55e;font-weight:700}}
.fail{{color:#ef4444;font-weight:700}}
.conclusion{{font-size:18px;margin-top:8px}}
tr:hover{{background:#fafafa}}
.log-path{{font-size:11px;max-width:200px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;font-family:monospace}}
</style>
</head>
<body>
<div class="container">
<h1>WiFi 异常自愈能力测试报告</h1>
<p class="meta">测试时间: {self.start_time} &nbsp;|&nbsp; 设备: {self.device_info}</p>
<div class="cards">
  <div class="card blue"><div class="num">{s['total']}</div><div class="label">总循环</div></div>
  <div class="card green"><div class="num">{s['m1_recovered']}</div><div class="label">方案一恢复</div></div>
  <div class="card orange"><div class="num">{s['degraded_rescued']}</div><div class="label">方案二救回</div></div>
  <div class="card {'green' if rate >= 95 else 'red'}"><div class="num">{rate:.1f}%</div><div class="label">综合成功率</div></div>
</div>
<div class="section">
<h2>降级链统计</h2>
<table>
<tr><th>指标</th><th>数值</th><th>说明</th></tr>
<tr><td>方案一成功率</td><td>{s['m1_rate']:.1f}%</td><td>开关WiFi即可恢复的比例</td></tr>
<tr><td>降级次数</td><td>{s['degraded']}</td><td>方案一失败，触发方案二</td></tr>
<tr><td>方案二救回</td><td>{s['degraded_rescued']} ({s['degraded_rescue_rate']:.1f}%)</td><td>飞行模式救回的比例</td></tr>
<tr><td>双双失败</td><td>{s['both_fail']}</td><td>两种方案均无法恢复</td></tr>
</table>
</div>
<div class="section">
<h2>方案统计</h2>
<table>
<tr><th>方案</th><th>执行次数</th><th>恢复成功</th><th>成功率</th><th>平均耗时</th><th>平均恢复率</th><th>开关正常</th><th>连接成功</th></tr>
<tr><td>方案一（开关WiFi）</td><td>{s['total']}</td><td>{s['m1_recovered']}/{s['total']}</td><td>{s['m1_rate']:.1f}%</td><td>{s['m1_avg_ms'] / 1000:.1f}s</td><td>{s['m1_avg_pct']:.0f}%</td><td>{s['m1_toggle_ok']}/{s['total']}</td><td>{s['m1_connect_ok']}/{s['total']}</td></tr>
<tr><td>方案二（飞行模式）</td><td>{s['m2_attempted']}</td><td>{s['m2_recovered']}/{s['m2_attempted']}</td><td>{s['m2_rate']:.1f}%</td><td>{s['m2_avg_ms'] / 1000:.1f}s</td><td>{s['m2_avg_pct']:.0f}%</td><td>{s['m2_toggle_ok']}/{s['m2_attempted']}</td><td>{s['m2_connect_ok']}/{s['m2_attempted']}</td></tr>
</table>
</div>
<div class="section">
<h2>故障类型统计</h2>
<table>
<tr><th>故障类型</th><th>出现次数</th><th>方案一成功</th><th>方案二救回</th><th>综合成功率</th></tr>
{fault_rows}
</table>
</div>
<div class="section">
<h2>全部记录</h2>
<table>
<tr><th>轮次</th><th>开始时间</th><th>开机完成</th><th>WiFi就绪</th><th>初始化耗时</th><th>列表为空</th><th>连接成功</th><th>本轮成功</th><th>失败类型</th><th>失败详情</th><th>重启OK</th><th>日志路径</th></tr>
{rows}
</table>
</div>
<div class="section">
<h2>结论</h2>
<p class="conclusion">{conclusion}</p>
</div>
</div>
</body>
</html>"""

        with open(path, "w", encoding="utf-8") as f:
            f.write(html)

    def save_excel(self, path: str):
        """生成 Excel 报告，包含汇总和逐轮明细两个 Sheet。"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            print("  [WARN] openpyxl 未安装，跳过 Excel 报告 (pip install openpyxl)")
            return

        s = self._compute_stats()
        wb = openpyxl.Workbook()

        # ---- 样式 ----
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        pass_font = Font(bold=True, color="228B22")
        fail_font = Font(bold=True, color="DC143C")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        center_align = Alignment(horizontal="center", vertical="center")

        # ================================================================
        # Sheet 1: 汇总
        # ================================================================
        ws1 = wb.active
        ws1.title = "汇总"

        summary_data = [
            ["WiFi 异常自愈能力测试报告", "", ""],
            ["", "", ""],
            ["测试时间", self.start_time, ""],
            ["设备", self.device_info, ""],
            ["总循环", s["total"], ""],
            ["恢复策略", RECOVERY_MODES.get(self.recovery_mode, self.recovery_mode), ""],
            ["目标SSID", self.ssid or "(自动检测)", ""],
            ["", "", ""],
            ["方案一（开关WiFi）", f"恢复 {s['m1_recovered']}/{s['total']} ({s['m1_rate']:.1f}%)",
             f"平均耗时 {s['m1_avg_ms']/1000:.1f}s  平均恢复率 {s['m1_avg_pct']:.0f}%"],
            ["  开关可点击", f"{s['m1_toggle_ok']}/{s['total']}", ""],
            ["  连接成功", f"{s['m1_connect_ok']}/{s['total']}", ""],
            ["方案二（飞行模式）", f"执行 {s['m2_attempted']} 次  恢复 {s['m2_recovered']}/{s['m2_attempted']} ({s['m2_rate']:.1f}%)",
             f"平均耗时 {s['m2_avg_ms']/1000:.1f}s"],
            ["", "", ""],
        ]

        if self.recovery_mode == "degrade":
            summary_data.append([
                "降级链", f"方案一失败 {s['degraded']} 次 → 方案二救回 {s['degraded_rescued']} 次 ({s['degraded_rescue_rate']:.1f}%)",
                f"双双失败: {s['both_fail']}",
            ])
        else:
            both_m1_ok = sum(1 for r in self.records if r.m1_recovered)
            both_m2_ok = sum(1 for r in self.records if r.m2_recovered)
            both_ok = sum(1 for r in self.records if r.m1_recovered and r.m2_recovered)
            summary_data.append([
                "全测模式", f"M1成功={both_m1_ok}  M2成功={both_m2_ok}  两者都成功={both_ok}",
                f"双双失败={s['both_fail']}",
            ])

        summary_data += [
            ["", "", ""],
            ["综合成功率", f"{s['overall_rate']:.1f}%",
             "PASS" if s['overall_rate'] >= 95 else "FAIL"],
        ]

        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, val in enumerate(row_data, 1):
                cell = ws1.cell(row=row_idx, column=col_idx, value=val)
                if row_idx == 1:
                    cell.font = Font(bold=True, size=14)
                elif row_idx in (9, 13):
                    cell.font = Font(bold=True, size=11)

        # 结论行
        last_row = len(summary_data)
        result_cell = ws1.cell(row=last_row, column=3)
        if s['overall_rate'] >= 95:
            result_cell.font = pass_font
        else:
            result_cell.font = fail_font

        ws1.column_dimensions['A'].width = 22
        ws1.column_dimensions['B'].width = 48
        ws1.column_dimensions['C'].width = 38

        # ================================================================
        # Sheet 2: 逐轮明细
        # ================================================================
        ws2 = wb.create_sheet(title="逐轮明细")

        headers = [
            "轮次", "开始时间", "开机完成", "WiFi就绪", "初始化耗时(s)",
            "列表为空", "连接成功", "本轮成功", "失败类型", "失败详情",
            "重启OK", "日志路径",
        ]
        for col_idx, h in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col_idx, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        for row_idx, r in enumerate(self.records, 2):
            scan_empty = "Y" if r.scan_empty else "N"
            final_conn = "OK" if r.final_connect_ok else ("-" if not self.ssid else "FAIL")
            success = "OK" if r.success else "FAIL"
            reboot_ok = "OK" if r.reboot_ok else "FAIL"
            fail_type = r.failure_type or "-"
            fail_detail = r.failure_detail or "-"
            log_path = r.log_path or "-"

            row_data = [
                r.cycle, r.timestamp[:19], r.boot_completed_time, r.wifi_ready_time,
                round(r.wifi_init_duration_s, 1), scan_empty, final_conn, success,
                fail_type, fail_detail, reboot_ok, log_path,
            ]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                cell.alignment = center_align
                if col_idx == 12:  # 日志路径左对齐
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                # 失败行标红
                if not r.success:
                    if col_idx == 8:  # 本轮成功列
                        cell.font = fail_font

        # 列宽
        col_widths = [6, 20, 10, 10, 12, 8, 8, 8, 12, 28, 8, 36]
        for i, w in enumerate(col_widths, 1):
            ws2.column_dimensions[get_column_letter(i)].width = w

        # 冻结首行
        ws2.freeze_panes = "A2"

        # 自动筛选
        ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(self.records) + 1}"

        wb.save(path)


# ============================================================
# CLI 入口
# ============================================================

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="WiFi 异常自愈能力测试 — 每轮重启设备，验证开机后 WiFi 恢复能力",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  一键启动: 双击 run.bat 或 命令行输入 run.bat
  python wifi_self_healing_test.py                     # 交互式配置
  python wifi_self_healing_test.py --cycles 50         # 50轮，自动检测WiFi
  python wifi_self_healing_test.py --cycles 100 --ssid "MyWiFi" --password "12345678"
  python wifi_self_healing_test.py --cycles 20 --boot-wait 45 --output ./results
        """,
    )
    parser.add_argument("--cycles", type=int, default=DEFAULT_CYCLES,
                        help=f"重启测试次数（默认: {DEFAULT_CYCLES}）")
    parser.add_argument("--boot-wait", type=int, default=DEFAULT_BOOT_WAIT,
                        help=f"启动完成后额外等待秒数，等待 WiFi 初始化（默认: {DEFAULT_BOOT_WAIT}）")
    parser.add_argument("--ssid", type=str, default=None,
                        help="目标 WiFi SSID，用于连接验证（不填则自动检测当前连接）")
    parser.add_argument("--password", type=str, default=None,
                        help="目标 WiFi 密码（不填则从设备自动读取已保存密码）")
    parser.add_argument("--no-auto-detect", action="store_true", default=False,
                        help="禁用自动检测WiFi凭据（需手动指定 --ssid 和 --password）")
    parser.add_argument("--output", type=str, default=DEFAULT_OUTPUT,
                        help=f"报告输出目录（默认: {DEFAULT_OUTPUT}）")
    parser.add_argument("--settings-intent", type=str, default=DEFAULT_SETTINGS_INTENT,
                        help=f"uiautomator2 降级时的设置页 Intent（默认: {DEFAULT_SETTINGS_INTENT}）")
    parser.add_argument("--recovery-mode", type=str, default=DEFAULT_RECOVERY_MODE,
                        choices=list(RECOVERY_MODES.keys()),
                        help=f"恢复策略: degrade(降级链，方案一失败再测方案二) / both(方案一二都测) （默认: {DEFAULT_RECOVERY_MODE}）")

    args = parser.parse_args()
    if args.cycles <= 0:
        parser.error("--cycles 必须为正整数")
    if args.password and not args.ssid:
        parser.error("--password 需要配合 --ssid 使用")
    return args


def interactive_config() -> argparse.Namespace:
    """交互式配置界面。"""
    print()
    print("=" * 50)
    print("   WiFi 异常自愈能力测试 — 配置")
    print("=" * 50)
    print("按 Enter 使用默认值，输入新值覆盖\n")

    def ask(label: str, default, cast=str):
        val = input(f"  {label} [{default}]: ").strip()
        if not val:
            return default
        try:
            return cast(val)
        except ValueError:
            print(f"    输入无效，使用默认值: {default}")
            return default

    cycles = ask("重启测试次数", DEFAULT_CYCLES, int)
    boot_wait = ask("开机启动等待秒数", DEFAULT_BOOT_WAIT, int)
    output = ask("报告输出目录", DEFAULT_OUTPUT)

    print()
    print("  --- 连接验证（可选）---")
    print("  留空 = 自动检测当前连接的 WiFi SSID 和密码")
    print("  输入 SSID = 手动指定 WiFi")
    ssid = ask("目标 WiFi SSID [自动检测]: ", "")
    password = ""
    if ssid:
        password = ask("  目标 WiFi 密码", "")

    print()
    print("  --- 恢复策略 ---")
    print("  降级链: 方案一失败后才测方案二")
    print("  全测:   方案一和方案二都测（同一故障）")
    test_both = input("  方案一和方案二都测？[y/N]: ").strip().lower()
    recovery_mode = "both" if test_both == "y" else DEFAULT_RECOVERY_MODE

    print()
    print("  --- 高级选项 ---")
    print("  设置页面路径用于 uiautomator2 降级方案，一般无需更改")
    print("  AOSP 原生: android.settings.WIFI_SETTINGS")
    print("  国内 ROM : android.settings.WIFI_SETTINGS (多数通用)")
    settings_intent = ask("设置页面 Intent", DEFAULT_SETTINGS_INTENT)

    print()
    print("-" * 50)
    print(f"  重启次数:      {cycles}")
    print(f"  启动等待:      {boot_wait}s")
    print(f"  报告目录:      {output}")
    print(f"  目标 WiFi:     {ssid if ssid else '(自动检测)'}")
    print(f"  设置页 Intent: {settings_intent}")
    print("-" * 50)

    confirm = input("\n  确认以上配置? [Y/n]: ").strip().lower()
    if confirm and confirm != "y":
        print("  已取消。")
        sys.exit(0)

    # 构造成和 argparse 一样的结果
    return argparse.Namespace(
        cycles=cycles,
        boot_wait=boot_wait,
        ssid=ssid if ssid else None,
        password=password if password else None,
        output=output,
        settings_intent=settings_intent,
        recovery_mode=recovery_mode,
        no_auto_detect=(ssid != ""),  # 手动指定了ssid则视为手动模式
    )


def _env_checks():
    """启动前环境检查：Python / ADB / 设备连接。失败则退出。"""
    print()
    print("=" * 55)
    print("   WiFi 异常自愈能力测试 — 环境检查")
    print("=" * 55)

    # Python 版本
    py_ver = f"{sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro}"
    print(f"   Python {py_ver}  ✓")

    # ADB 可用性
    try:
        proc = subprocess.run(
            ["adb", "version"], capture_output=True, text=True, timeout=5,
            encoding="utf-8", errors="replace",
        )
        if proc.returncode == 0:
            ver_line = proc.stdout.strip().split("\n")[0] if proc.stdout else "OK"
            print(f"   ADB {ver_line.split('version')[0].strip() if 'version' in ver_line else '已安装'}  ✓")
        else:
            print("   ADB [FAIL] — 未找到 ADB，请确认已安装并加入 PATH")
            print("     下载: https://developer.android.com/studio/releases/platform-tools")
            sys.exit(1)
    except FileNotFoundError:
        print("   ADB [FAIL] — 未找到 ADB，请确认已安装并加入 PATH")
        sys.exit(1)

    # 设备连接
    try:
        proc = subprocess.run(
            ["adb", "devices"], capture_output=True, text=True, timeout=10,
            encoding="utf-8", errors="replace",
        )
        lines = [l.strip() for l in proc.stdout.split("\n") if l.strip()]
        devices = [l for l in lines[1:] if l and "\tdevice" in l] if len(lines) > 1 else []
        if devices:
            print(f"   设备 {devices[0].split()[0]}  ✓")
        else:
            print("   设备 [FAIL] — 未检测到已连接的设备")
            print("   当前 adb devices 输出:")
            for l in lines:
                print(f"     {l}")
            print("   请确认: USB已连接 / USB调试已开启 / 已授权此电脑")
            sys.exit(1)
    except Exception:
        print("   设备 [FAIL] — 无法执行 adb devices")
        sys.exit(1)

    # uiautomator2
    try:
        import uiautomator2
        print("   uiautomator2 已安装  ✓")
    except ImportError:
        print("   uiautomator2 未安装 (可选，不影响主功能)")

    print("=" * 55)


def main():
    # 环境检查（无参数时先检查，有参数时跳过直接进测试）
    if len(sys.argv) <= 1:
        _env_checks()

    # 如果命令行传了参数则用参数，否则交互式配置
    if len(sys.argv) > 1:
        args = parse_args()
    else:
        args = interactive_config()

    runner = TestRunner(
        cycles=args.cycles,
        boot_wait=args.boot_wait,
        ssid=args.ssid,
        password=args.password,
        auto_detect=not args.no_auto_detect,
        settings_intent=args.settings_intent,
        recovery_mode=args.recovery_mode,
    )
    records = runner.run()

    reporter = ReportGenerator(
        records=records,
        device_info=runner.device_info,
        start_time=runner.start_time,
        output_dir=args.output,
        ssid=args.ssid,
        recovery_mode=args.recovery_mode,
    )
    reporter.generate_all()


if __name__ == "__main__":
    main()
